import requests
import urllib.request
import openpyxl
from pyvirtualdisplay import Display

wk = openpyxl.load_workbook("/home/walterybarra/Documentos/Test-evidencia/PaginaBlock.xlsx")
sh = wk["Reporte"]
tb = openpyxl.Workbook()
hj = tb.active
hj.title="Reporte"
f = open ('/home/walterybarra/Documentos/Test-evidencia/Reporte_responce.txt','w')
f.write('Log de URL'+"\n")
display = Display(visible=0, size=(1200, 600))
display.start()
filas = sh.max_row
columnas = sh.max_column
count=1
#--------------------

#---------------------
print("======================= I n i c i o ================================")
# Columnas de  usuarios
for j in range(1, filas + 1):
    c = sh.cell(j, 1)
    d = sh.cell(j, 2)
    url=c.value
    url2 = "https://"+str(c.value)
    status_code_url = requests.get(url2).status_code
    cadena = str(c.value)[-5:]
    if status_code_url == 404:
        print("404 - "+str(count)+"-"+cadena+"-"+str(c.value))
        f.write("404 - "+str(count)+"--"+str(c.value)+"\n ")
        hj['A' + str(j)].value = "URL Con 404" #tabla
        hj['B' + str(j)].value = cadena  # tabla
        hj['C' + str(j)].value = str(c.value)  # tabla
    elif status_code_url == 200:

        f.write("Error"+" Fila "+str(count)+"-> "+str(c.value)+"\n ")
        hj['A' + str(j)].value = "URL Activo" #codigo
        hj['B' + str(j)].value = cadena  # numero de fila
        hj['C' + str(j)].value = str(c.value)  # url

        #----------------------------------- -----------------------
        r = requests.get(url2)
        hj['D' + str(j)].value = r.url  # redireccion
        print("200 - " + str(count) + "-" + cadena + "-" + str(c.value)+" Redireccion "+r.url)
        # print(r.url)
        #-----------------------------------------------------------
    count = count + 1

print("======================= F i n ================================")

tb.save("/home/walterybarra/Documentos/Test-evidencia/reporte.xlsx")
f.close()
